import os
import logging
import csv
from dotenv import load_dotenv
import psycopg2
from typing import List, Dict, Any, Tuple, Optional
import uuid
from datetime import datetime
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()

# Database configurations
DB_CONFIGS = {
    "sc_ep": {
        "host": os.getenv("SC_EP_PG_DB_HOST"),
        "user": os.getenv("SC_EP_PG_DB_USER"),
        "password": os.getenv("SC_EP_PG_DB_PASSWORD"),
        "database": os.getenv("SC_EP_PG_DB_NAME"),
        "port": 5432
    }
}

# Constants
DEFAULT_ENGINEER_ID = "895"
DEFAULT_ASSIGNED_BY = "895"
DEFAULT_CREATED_BY = 895
DEFAULT_APPROVED_BY = 895
DEFAULT_SUBMITTED_BY = 895
ASSIGNMENT_TYPE = "INSTALLATION"
JSF_STATUS = "COMPLETED"
JSF_TYPE = "INSTALLATION"
OUTCOME_REASON = "Self_Installation"
ENGINEER_RECOMMENDATION = "Installed"

# Global customer to account mapping
# Key: customer_id (matches customer_id field in premises table)
# Value: List of account_ids that belong to this customer
CUSTOMER_ACCOUNT_MAPPING = {
    1106: [16497],
}

class AccountProcessingError(Exception):
    """Custom exception for account processing errors"""
    pass

def establish_pg_db_connection(db_config_name: str) -> psycopg2.extensions.connection:
    """Establish a connection to the PostgreSQL database."""
    try:
        connection = psycopg2.connect(**DB_CONFIGS[db_config_name])
        logger.info(f"Successfully connected to {db_config_name} database")
        return connection
    except psycopg2.Error as e:
        logger.error(f"Error connecting to {db_config_name} PostgreSQL: {e}")
        raise

def validate_customer_premise_details(connection: psycopg2.extensions.connection, 
                                      customer_account_mapping: Dict[int, List[int]]) -> Tuple[Dict[int, uuid.UUID], List[int]]:
    """
    Validate that customers exist in premises table using customer_id field.
    If customer exists, all their accounts from the mapping are considered valid.
    
    Returns:
        Tuple of (valid_account_premises_map, invalid_accounts)
        - valid_account_premises_map: Dict mapping account_id to premises_id (as UUID) for valid customers
        - invalid_accounts: List of account_ids whose customers don't exist in premises table
    """
    try:
        cursor = connection.cursor()
        valid_account_premises_map = {}
        invalid_accounts = []
        
        if not customer_account_mapping:
            raise AccountProcessingError("No customer-account mapping provided")
        
        for customer_id, account_ids in customer_account_mapping.items():
            logger.info(f"Validating customer {customer_id}")
            
            # Check if customer exists in premises table using customer_id field
            customer_query = """
                SELECT id 
                FROM public.premises 
                WHERE customer_id = %s
                LIMIT 1
            """
            
            cursor.execute(customer_query, (customer_id,))
            customer_result = cursor.fetchone()
            
            if not customer_result:
                logger.error(f"Customer {customer_id} not found in premises table")
                invalid_accounts.extend(account_ids)
                continue
            
            premises_id = customer_result[0]  # This is already a UUID object from psycopg2
            logger.info(f"Customer {customer_id} found with premises_id {premises_id}")
            
            # Since customer exists, all their accounts are considered valid
            for account_id in account_ids:
                valid_account_premises_map[account_id] = premises_id
                logger.info(f"Account {account_id} marked as valid for customer {customer_id}")
        
        logger.info(f"Validation complete: {len(valid_account_premises_map)} valid accounts, {len(invalid_accounts)} invalid accounts")
        return valid_account_premises_map, invalid_accounts
        
    except psycopg2.Error as e:
        logger.error(f"Database error while validating customers: {e}")
        raise
    finally:
        cursor.close()

def filter_accounts_without_jsf(connection: psycopg2.extensions.connection,
                               account_premises_map: Dict[int, uuid.UUID]) -> Dict[int, uuid.UUID]:
    """
    Filter out accounts that already have job satisfaction forms with the specific criteria.
    Only returns accounts that need new JSF workflow (assignment → schedule → JSF).
    
    Returns:
        Dictionary mapping account_id to premises_id for accounts without matching JSF
    """
    try:
        cursor = connection.cursor()
        account_ids = list(account_premises_map.keys())
        
        if not account_ids:
            return {}
            
        placeholders = ','.join(['%s'] * len(account_ids))
        
        # Check for existing JSF with specific criteria by joining through schedules and assignments
        query = f"""
            SELECT DISTINCT a.account_id 
            FROM public.assignments a
            JOIN public.schedules s ON a.id = s.assignment_id
            JOIN public.job_satisfaction_form jsf ON s.id = jsf.schedule_id
            WHERE a.account_id IN ({placeholders})
            AND jsf.jsf_type = %s
            AND jsf.outcome_reason = %s
            AND jsf.engineer_recommendation = %s
        """
        
        cursor.execute(query, account_ids + [JSF_TYPE, OUTCOME_REASON, ENGINEER_RECOMMENDATION])
        existing_jsf_accounts = [row[0] for row in cursor.fetchall()]
        
        # Filter out accounts that already have the specific JSF
        accounts_without_jsf = {
            acc_id: premises_id 
            for acc_id, premises_id in account_premises_map.items() 
            if acc_id not in existing_jsf_accounts
        }
        
        if existing_jsf_accounts:
            logger.info(f"Skipping {len(existing_jsf_accounts)} accounts with existing JSF (INSTALLATION/Self_Installation/Installed): {existing_jsf_accounts}")
        
        logger.info(f"Found {len(accounts_without_jsf)} accounts without matching JSF")
        return accounts_without_jsf
        
    except psycopg2.Error as e:
        logger.error(f"Error checking existing job satisfaction forms: {e}")
        raise
    finally:
        cursor.close()

def create_assignments_batch(connection: psycopg2.extensions.connection, 
                             account_premises_map: Dict[int, uuid.UUID]) -> Dict[int, uuid.UUID]:
    """
    Create assignment records for accounts using batch processing.
    
    Returns:
        Dictionary mapping account_id to assignment_id (UUID)
    """
    try:
        cursor = connection.cursor()
        account_assignment_map = {}
        
        if not account_premises_map:
            logger.info("No accounts to create assignments for")
            return account_assignment_map
        
        insert_query = """
            INSERT INTO public.assignments (
                meta, created_at, updated_at, created_by, updated_by,
                is_active, premises_id, engineer_id, assignment_type,
                assigned_by, assignment_date, account_id, ticket_id,
                ticket_number, number, comment
            ) VALUES (
                %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP,
                %s, %s, %s, %s, %s, %s, %s,
                CURRENT_TIMESTAMP, %s, %s, %s, %s, %s
            )
            RETURNING id, account_id
        """
        
        for account_id, premises_id in account_premises_map.items():
            cursor.execute(insert_query, (
                '{}',                      # meta
                DEFAULT_CREATED_BY,        # created_by
                DEFAULT_CREATED_BY,        # updated_by
                False,                     # is_active
                premises_id,               # premises_id (UUID)
                DEFAULT_ENGINEER_ID,       # engineer_id (string)
                ASSIGNMENT_TYPE,           # assignment_type
                DEFAULT_ASSIGNED_BY,       # assigned_by (string)
                account_id,                # account_id
                None, None, None, None     # ticket_id, ticket_number, number, comment
            ))
            result = cursor.fetchone()
            if result:
                assignment_id, returned_account_id = result
                account_assignment_map[returned_account_id] = assignment_id  # Keep as UUID
        
        logger.info(f"Successfully created {len(account_assignment_map)} assignments")
        
        if len(account_assignment_map) == 0 and len(account_premises_map) > 0:
            raise AccountProcessingError("Failed to create assignments: no rows affected")
        
        return account_assignment_map
        
    except psycopg2.Error as e:
        logger.error(f"Error creating assignments: {e}")
        raise
    finally:
        cursor.close()

def create_schedules_batch(connection: psycopg2.extensions.connection,
                           account_assignment_map: Dict[int, uuid.UUID]) -> Dict[int, uuid.UUID]:
    """
    Create schedule records for each assignment using batch processing.
    
    Returns:
        Dictionary mapping account_id to schedule_id (UUID)
    """
    try:
        cursor = connection.cursor()
        account_schedule_map = {}
        
        if not account_assignment_map:
            logger.info("No assignments to create schedules for")
            return account_schedule_map
        
        insert_query = """
            INSERT INTO public.schedules (
                meta, created_at, updated_at, created_by, updated_by,
                is_active, assignment_id, scheduled_by, scheduled_date,
                completed_date
            ) VALUES (
                %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP,
                %s, %s, %s, %s, %s, CURRENT_DATE, CURRENT_DATE
            )
            RETURNING id, assignment_id
        """
        
        assignment_to_account = {v: k for k, v in account_assignment_map.items()}
        
        for account_id, assignment_id in account_assignment_map.items():
            cursor.execute(insert_query, (
                '{}',                    # meta
                DEFAULT_CREATED_BY,      # created_by
                DEFAULT_CREATED_BY,      # updated_by
                False,                   # is_active
                assignment_id,           # assignment_id (UUID)
                DEFAULT_ASSIGNED_BY      # scheduled_by (string)
            ))
            result = cursor.fetchone()
            if result:
                schedule_id, returned_assignment_id = result
                account_id_for_schedule = assignment_to_account[returned_assignment_id]
                account_schedule_map[account_id_for_schedule] = schedule_id  # Keep as UUID
        
        logger.info(f"Successfully created {len(account_schedule_map)} schedules")
        return account_schedule_map
        
    except psycopg2.Error as e:
        logger.error(f"Error creating schedules: {e}")
        raise
    finally:
        cursor.close()

def create_job_satisfaction_forms_batch(connection: psycopg2.extensions.connection,
                                        account_schedule_map: Dict[int, uuid.UUID]) -> Dict[int, uuid.UUID]:
    """
    Create job satisfaction form records for each schedule using batch processing.
    
    Returns:
        Dictionary mapping account_id to jsf_id (UUID)
    """
    try:
        cursor = connection.cursor()
        account_jsf_map = {}
        
        if not account_schedule_map:
            logger.info("No schedules to create JSFs for")
            return account_jsf_map
        
        insert_query = """
            INSERT INTO public.job_satisfaction_form (
                meta, created_at, updated_at, created_by, updated_by,
                is_active, jsf_status, schedule_id, jsf_type, completed_date,
                device_id, device_status, casual_pay, costings, outcome_reason,
                jsf_start_time, jsf_end_time, engineer_recommendation, device_image,
                product_type, comment, approval_date, approved_by, submission_date,
                submitted_by
            ) VALUES (
                %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP,
                %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP,
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s,
                %s, %s, CURRENT_TIMESTAMP, %s, CURRENT_TIMESTAMP,
                %s
            )
            RETURNING id, schedule_id
        """
        
        schedule_to_account = {v: k for k, v in account_schedule_map.items()}
        
        for account_id, schedule_id in account_schedule_map.items():
            cursor.execute(insert_query, (
                '{}',                          # meta
                DEFAULT_CREATED_BY,            # created_by
                DEFAULT_CREATED_BY,            # updated_by
                False,                         # is_active
                JSF_STATUS,                    # jsf_status
                schedule_id,                   # schedule_id (UUID)
                JSF_TYPE,                      # jsf_type
                None, None, None, None,        # device_id, device_status, casual_pay, costings
                OUTCOME_REASON,                # outcome_reason
                None, None,                    # jsf_start_time, jsf_end_time
                ENGINEER_RECOMMENDATION,       # engineer_recommendation
                None, None, None,              # device_image, product_type, comment
                DEFAULT_APPROVED_BY,           # approved_by
                DEFAULT_SUBMITTED_BY           # submitted_by
            ))
            result = cursor.fetchone()
            if result:
                jsf_id, returned_schedule_id = result
                account_id_for_jsf = schedule_to_account[returned_schedule_id]
                account_jsf_map[account_id_for_jsf] = jsf_id  # Keep as UUID
        
        logger.info(f"Successfully created {len(account_jsf_map)} job satisfaction forms")
        return account_jsf_map
        
    except psycopg2.Error as e:
        logger.error(f"Error creating job satisfaction forms: {e}")
        raise
    finally:
        cursor.close()

def generate_report(customer_account_mapping: Dict[int, List[int]],
                   processed_accounts: List[int], 
                   skipped_accounts: List[int],
                   invalid_accounts: List[int],
                   account_premises_map: Dict[int, uuid.UUID],
                   account_assignment_map: Dict[int, uuid.UUID],
                   account_schedule_map: Dict[int, uuid.UUID],
                   account_jsf_map: Dict[int, uuid.UUID]) -> str:
    """
    Generate Excel report with multiple sheets for accounts processing results.
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"assignment_report_{timestamp}.xlsx"
    
    try:
        workbook = openpyxl.Workbook()
        
        # Sheet 1: Processing Summary
        sheet1: Worksheet = workbook.active
        sheet1.title = "Processing Summary"
        sheet1.append(['Category', 'Count', 'Timestamp'])
        sheet1.append(['Total Customers', len(customer_account_mapping), datetime.now().isoformat()])
        sheet1.append(['Total Accounts', sum(len(accs) for accs in customer_account_mapping.values()), datetime.now().isoformat()])
        sheet1.append(['Successfully Processed', len(processed_accounts), datetime.now().isoformat()])
        sheet1.append(['Skipped (Existing JSF)', len(skipped_accounts), datetime.now().isoformat()])
        sheet1.append(['Invalid Accounts', len(invalid_accounts), datetime.now().isoformat()])
        
        # Sheet 2: Successfully Processed
        sheet2: Worksheet = workbook.create_sheet("Successfully Processed")
        sheet2.append(['Account ID', 'Premises ID', 'Assignment ID', 'Schedule ID', 'JSF ID', 'Timestamp'])
        for account_id in processed_accounts:
            sheet2.append([
                account_id,
                str(account_premises_map.get(account_id, 'N/A')),
                str(account_assignment_map.get(account_id, 'N/A')),
                str(account_schedule_map.get(account_id, 'N/A')),
                str(account_jsf_map.get(account_id, 'N/A')),
                datetime.now().isoformat()
            ])
        
        # Sheet 3: Skipped Accounts
        sheet3: Worksheet = workbook.create_sheet("Skipped Accounts")
        sheet3.append(['Account ID', 'Premises ID', 'Reason', 'Timestamp'])
        for account_id in skipped_accounts:
            sheet3.append([
                account_id,
                str(account_premises_map.get(account_id, 'N/A')),
                'JSF already exists (INSTALLATION/Self_Installation/Installed)',
                datetime.now().isoformat()
            ])
        
        # Sheet 4: Invalid Accounts
        sheet4: Worksheet = workbook.create_sheet("Invalid Accounts")
        sheet4.append(['Account ID', 'Reason', 'Timestamp'])
        for account_id in invalid_accounts:
            sheet4.append([
                account_id,
                'Customer not found in premises table',
                datetime.now().isoformat()
            ])
        
        workbook.save(filename)
        logger.info(f"Excel report generated: {filename}")
        return filename
        
    except Exception as e:
        logger.error(f"Error generating Excel report: {e}")
        return ""

def process_account_assignments(customer_account_mapping: Dict[int, List[int]]) -> Tuple[int, List[int], List[int], str]:
    """
    Main function to process account assignments with customer-account mapping approach.
    Returns: (success_count, skipped_accounts, invalid_accounts, report_filename)
    """
    connection = None
    try:
        # Establish database connection
        connection = establish_pg_db_connection("sc_ep")
        connection.autocommit = False  # Ensure we're in transaction mode
        
        logger.info(f"Processing {len(customer_account_mapping)} customers with {sum(len(accs) for accs in customer_account_mapping.values())} total accounts")
        
        # Step 1: Validate customers exist in premises table
        valid_account_premises_map, invalid_accounts = validate_customer_premise_details(connection, customer_account_mapping)
        
        if not valid_account_premises_map:
            logger.info("No valid accounts found to process")
            report_filename = generate_report(customer_account_mapping, [], [], invalid_accounts, 
                                             {}, {}, {}, {})
            connection.commit()
            return 0, [], invalid_accounts, report_filename
        
        # Step 2: Filter out accounts that already have matching JSF
        accounts_needing_jsf = filter_accounts_without_jsf(connection, valid_account_premises_map)
        
        # Determine skipped accounts (those with existing matching JSF)
        skipped_accounts = [acc_id for acc_id in valid_account_premises_map.keys() 
                           if acc_id not in accounts_needing_jsf]
        
        if not accounts_needing_jsf:
            logger.info("No new JSF workflows needed - all valid accounts already have matching JSF")
            report_filename = generate_report(customer_account_mapping, [], skipped_accounts, invalid_accounts, 
                                             valid_account_premises_map, {}, {}, {})
            connection.commit()
            return 0, skipped_accounts, invalid_accounts, report_filename
        
        # Step 3: Create assignments
        account_assignment_map = create_assignments_batch(connection, accounts_needing_jsf)
        
        # Step 4: Create schedules
        account_schedule_map = create_schedules_batch(connection, account_assignment_map)
        
        # Step 5: Create job satisfaction forms
        account_jsf_map = create_job_satisfaction_forms_batch(connection, account_schedule_map)
        
        # Commit all transactions
        connection.commit()
        
        # Generate report
        successful_accounts = list(account_assignment_map.keys())
        report_filename = generate_report(
            customer_account_mapping,
            successful_accounts, 
            skipped_accounts, 
            invalid_accounts,
            valid_account_premises_map, 
            account_assignment_map, 
            account_schedule_map,
            account_jsf_map
        )
        
        logger.info(f"Processing completed successfully:")
        logger.info(f"  - {len(successful_accounts)} new workflows created (assignment → schedule → JSF)")
        logger.info(f"  - {len(skipped_accounts)} accounts skipped (existing matching JSF)")
        logger.info(f"  - {len(invalid_accounts)} invalid accounts")
        logger.info(f"  - Report file: {report_filename}")
        
        return len(successful_accounts), skipped_accounts, invalid_accounts, report_filename
        
    except Exception as e:
        logger.error(f"Error in process_account_assignments: {e}")
        if connection:
            connection.rollback()
        # Get all account IDs from mapping for error reporting
        all_account_ids = [acc_id for acc_list in customer_account_mapping.values() for acc_id in acc_list]
        return 0, [], all_account_ids, ""
        
    finally:
        if connection:
            connection.close()

def main(customer_account_mapping: Optional[Dict[int, List[int]]] = None) -> Tuple[int, List[int], List[int], str]:
    """Main function to execute the assignment creation process."""
    try:
        logger.info("=== Starting Assignment Creation Process ===")
        
        mapping_to_process = customer_account_mapping or CUSTOMER_ACCOUNT_MAPPING
        
        if not mapping_to_process:
            logger.error("No customer-account mapping to process!")
            return 0, [], [], ""
        
        # Process assignments
        success_count, skipped_accounts, invalid_accounts, report_file = process_account_assignments(mapping_to_process)
        
        logger.info("=== Process Completed ===")
        return success_count, skipped_accounts, invalid_accounts, report_file
        
    except Exception as e:
        logger.error(f"Error in main process: {e}")
        raise

if __name__ == "__main__":
    # Run the process with the global mapping
    main()