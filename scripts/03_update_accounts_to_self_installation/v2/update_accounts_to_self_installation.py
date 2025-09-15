import os
import logging
import csv
from dotenv import load_dotenv
import psycopg2
from typing import List, Dict, Any, Tuple, Optional
import uuid
from datetime import datetime
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()

# Database configurations
DB_CONFIGS = {
    "sc_ep": {
        "host": os.getenv("SC_EP_PG_DB_HOST"),
        "user": os.getenv("SC_EP_PG_DB_USER"),
        "password": os.getenv("SC_EP_PG_DB_PASSWORD"),
        "database": os.getenv("SC_EP_PG_DB_NAME"),
        "port": 5432
    }
}

# Constants
DEFAULT_ENGINEER_ID = "895"  # Changed to string to match character varying
DEFAULT_ASSIGNED_BY = "895"  # Changed to string to match character varying
DEFAULT_CREATED_BY = 895
ASSIGNMENT_TYPE = "INSTALLATION"

# Global account IDs
ACCOUNT_IDS = [
    49680
]

class AccountProcessingError(Exception):
    """Custom exception for account processing errors"""
    pass

def establish_pg_db_connection(db_config_name: str) -> psycopg2.extensions.connection:
    """Establish a connection to the PostgreSQL database."""
    try:
        connection = psycopg2.connect(**DB_CONFIGS[db_config_name])
        logger.info(f"Successfully connected to {db_config_name} database")
        return connection
    except psycopg2.Error as e:
        logger.error(f"Error connecting to {db_config_name} PostgreSQL: {e}")
        raise

def validate_all_accounts_exist(connection: psycopg2.extensions.connection, 
                               account_ids: List[int]) -> Dict[int, uuid.UUID]:
    """
    Validate that ALL account IDs exist in the premises table's account_id array.
    If any account is missing, raises an exception to stop execution.
    
    Returns:
        Dictionary mapping account_id to premises_id
    """
    try:
        cursor = connection.cursor()
        
        if not account_ids:
            raise AccountProcessingError("No account IDs provided")
            
        # Use unnest to expand account_id array and filter to match input account_ids
        placeholders = ','.join(['%s'] * len(account_ids))
        query = f"""
            SELECT aid as account_id, p.id as premises_id 
            FROM public.premises p, unnest(p.account_id) as aid 
            WHERE p.account_id && ARRAY[{placeholders}]::integer[] 
            AND aid = ANY(ARRAY[{placeholders}]::integer[])
        """
        
        cursor.execute(query, account_ids + account_ids)  # Pass account_ids twice for both placeholders
        results = cursor.fetchall()
        
        # Map individual account_id (integer) to premises_id (uuid)
        account_premises_map = {row[0]: uuid.UUID(str(row[1])) for row in results}
        missing_accounts = [acc_id for acc_id in account_ids if acc_id not in account_premises_map]
        
        if missing_accounts:
            logger.error(f"Missing accounts in premises table: {missing_accounts}")
            raise AccountProcessingError(f"The following accounts do not exist in premises table: {missing_accounts}")
        
        logger.info(f"All {len(account_premises_map)} accounts found in premises table")
        return account_premises_map
        
    except psycopg2.Error as e:
        logger.error(f"Database error while validating accounts: {e}")
        raise
    finally:
        cursor.close()

def filter_accounts_without_assignments(connection: psycopg2.extensions.connection,
                                      account_premises_map: Dict[int, uuid.UUID]) -> Dict[int, uuid.UUID]:
    """
    Filter out accounts that already have assignments.
    Only returns accounts that need new assignments.
    
    Returns:
        Dictionary mapping account_id to premises_id for accounts without assignments
    """
    try:
        cursor = connection.cursor()
        account_ids = list(account_premises_map.keys())
        
        if not account_ids:
            return {}
            
        placeholders = ','.join(['%s'] * len(account_ids))
        query = f"""
            SELECT DISTINCT account_id 
            FROM public.assignments 
            WHERE account_id IN ({placeholders}) 
        """
        
        cursor.execute(query, account_ids)
        existing_assignments = [row[0] for row in cursor.fetchall()]
        
        # Filter out accounts that already have assignments
        accounts_without_assignments = {
            acc_id: premises_id 
            for acc_id, premises_id in account_premises_map.items() 
            if acc_id not in existing_assignments
        }
        
        if existing_assignments:
            logger.info(f"Skipping {len(existing_assignments)} accounts with existing assignments: {existing_assignments}")
        
        logger.info(f"Found {len(accounts_without_assignments)} accounts without assignments")
        return accounts_without_assignments
        
    except psycopg2.Error as e:
        logger.error(f"Error checking existing assignments: {e}")
        raise
    finally:
        cursor.close()

def create_assignments_batch(connection: psycopg2.extensions.connection, 
                           account_premises_map: Dict[int, uuid.UUID],
                           engineer_id: str = DEFAULT_ENGINEER_ID,
                           assigned_by: str = DEFAULT_ASSIGNED_BY,
                           created_by: int = DEFAULT_CREATED_BY) -> Dict[int, str]:
    """
    Create assignment records for accounts using batch processing.
    
    Returns:
        Dictionary mapping account_id to assignment_id
    """
    try:
        cursor = connection.cursor()
        account_assignment_map = {}
        
        if not account_premises_map:
            logger.info("No accounts to create assignments for")
            return account_assignment_map
        
        insert_query = """
            INSERT INTO public.assignments (
                meta, created_at, updated_at, created_by, updated_by,
                is_active, premises_id, engineer_id, assignment_type,
                assigned_by, assignment_date, account_id, ticket_id,
                ticket_number, number, comment
            ) VALUES (
                %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP,
                %s, %s, %s, %s, %s, %s, %s,
                CURRENT_TIMESTAMP, %s, %s, %s, %s, %s
            )
            RETURNING id, account_id
        """
        
        # Use execute() in a loop instead of executemany() to get RETURNING results
        for account_id, premises_id in account_premises_map.items():
            cursor.execute(insert_query, (
                '{}',                    # meta
                created_by,
                created_by,
                False,
                str(premises_id),        # Convert UUID to string
                engineer_id,             # string
                ASSIGNMENT_TYPE,
                assigned_by,             # string
                account_id,
                None, None, None, None
            ))
            result = cursor.fetchone()
            if result:
                assignment_id, returned_account_id = result
                account_assignment_map[returned_account_id] = str(assignment_id)
        
        row_count = len(account_assignment_map)
        logger.info(f"Inserted {row_count} rows into assignments table")
        
        if row_count == 0:
            logger.error("No rows inserted into assignments table")
            raise AccountProcessingError("Failed to insert assignments: no rows affected")
        
        logger.info(f"Successfully created {len(account_assignment_map)} assignments")
        return account_assignment_map
        
    except psycopg2.Error as e:
        logger.error(f"Error creating assignments: {e}")
        raise
    finally:
        cursor.close()

def create_schedules_batch(connection: psycopg2.extensions.connection,
                         account_assignment_map: Dict[int, str],
                         scheduled_by: str = DEFAULT_ASSIGNED_BY,
                         created_by: int = DEFAULT_CREATED_BY) -> Dict[int, str]:
    """
    Create schedule records for each assignment using batch processing.
    
    Returns:
        Dictionary mapping account_id to schedule_id
    """
    try:
        cursor = connection.cursor()
        account_schedule_map = {}
        
        if not account_assignment_map:
            logger.info("No assignments to create schedules for")
            return account_schedule_map
        
        insert_query = """
            INSERT INTO public.schedules (
                meta, created_at, updated_at, created_by, updated_by,
                is_active, assignment_id, scheduled_by, scheduled_date,
                completed_date
            ) VALUES (
                %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP,
                %s, %s, %s, %s, %s, CURRENT_DATE, CURRENT_DATE
            )
            RETURNING id, assignment_id
        """
        
        # Use execute() in a loop instead of executemany() to get RETURNING results
        assignment_to_account = {str(v): k for k, v in account_assignment_map.items()}
        
        for account_id, assignment_id in account_assignment_map.items():
            cursor.execute(insert_query, (
                '{}',  # meta
                created_by,
                created_by,
                False,
                assignment_id,  # Keep as string
                scheduled_by
            ))
            result = cursor.fetchone()
            if result:
                schedule_id, returned_assignment_id = result
                account_id_for_schedule = assignment_to_account[str(returned_assignment_id)]
                account_schedule_map[account_id_for_schedule] = str(schedule_id)
        
        logger.info(f"Successfully created {len(account_schedule_map)} schedules")
        return account_schedule_map
        
    except psycopg2.Error as e:
        logger.error(f"Error creating schedules: {e}")
        raise
    finally:
        cursor.close()

def create_job_satisfaction_forms_batch(connection: psycopg2.extensions.connection,
                                      account_schedule_map: Dict[int, str],
                                      created_by: int = DEFAULT_CREATED_BY,
                                      approved_by: int = 895,
                                      submitted_by: int = 895) -> Dict[int, str]:
    """
    Create job satisfaction form records for each schedule using batch processing.
    
    Returns:
        Dictionary mapping account_id to jsf_id
    """
    try:
        cursor = connection.cursor()
        account_jsf_map = {}
        
        if not account_schedule_map:
            logger.info("No schedules to create JSFs for")
            return account_jsf_map
        
        insert_query = """
            INSERT INTO public.job_satisfaction_form (
                meta, created_at, updated_at, created_by, updated_by,
                is_active, jsf_status, schedule_id, jsf_type, completed_date,
                device_id, device_status, casual_pay, costings, outcome_reason,
                jsf_start_time, jsf_end_time, engineer_recommendation, device_image,
                product_type, comment, approval_date, approved_by, submission_date,
                submitted_by
            ) VALUES (
                %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP,
                %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP,
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s,
                %s, %s, CURRENT_TIMESTAMP, %s, CURRENT_TIMESTAMP,
                %s
            )
            RETURNING id, schedule_id
        """
        
        # Use execute() in a loop instead of executemany() to get RETURNING results
        schedule_to_account = {str(v): k for k, v in account_schedule_map.items()}
        
        for account_id, schedule_id in account_schedule_map.items():
            cursor.execute(insert_query, (
                '{}',           # meta
                created_by,     # created_by
                created_by,     # updated_by
                False,           # is_active
                'COMPLETED',    # jsf_status
                schedule_id,    # Keep as string
                'INSTALLATION', # jsf_type
                None, None, None, None,   # device_id, device_status, casual_pay, costings
                'Self_Installation',  # outcome_reason
                None, None,            # jsf_start_time, jsf_end_time
                'Installed',           # engineer_recommendation
                None, None, None,      # device_image, product_type, comment
                approved_by,
                submitted_by
            ))
            result = cursor.fetchone()
            if result:
                jsf_id, returned_schedule_id = result
                account_id_for_jsf = schedule_to_account[str(returned_schedule_id)]
                account_jsf_map[account_id_for_jsf] = str(jsf_id)
        
        logger.info(f"Successfully created {len(account_jsf_map)} job satisfaction forms")
        return account_jsf_map
        
    except psycopg2.Error as e:
        logger.error(f"Error creating job satisfaction forms: {e}")
        raise
    finally:
        cursor.close()
        

        
def generate_report(account_ids: List[int],
                   processed_accounts: List[int], 
                   skipped_accounts: List[int],
                   missing_premises_accounts: List[int],
                   account_premises_map: Dict[int, uuid.UUID],
                   account_assignment_map: Dict[int, str],
                   account_schedule_map: Dict[int, str],
                   account_jsf_map: Dict[int, str]) -> str:
    """
    Generate Excel report with multiple sheets for accounts to update, missing premises, 
    assignments, schedules, and JSFs.
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"assignment_report_{timestamp}.xlsx"
    
    try:
        workbook = openpyxl.Workbook()
        
        # Sheet 1: Accounts to Update
        sheet1: Worksheet = workbook.active
        sheet1.title = "Accounts to Update"
        sheet1.append(['Account ID', 'Timestamp'])
        for account_id in account_ids:
            sheet1.append([account_id, datetime.now().isoformat()])
        
        # Sheet 2: Missing Premises
        sheet2: Worksheet = workbook.create_sheet("Missing Premises")
        sheet2.append(['Account ID', 'Status', 'Timestamp'])
        for account_id in missing_premises_accounts:
            sheet2.append([account_id, 'FAILED - Missing premises', datetime.now().isoformat()])
        
        # Sheet 3: Assignments
        sheet3: Worksheet = workbook.create_sheet("Assignments")
        sheet3.append(['Account ID', 'Status', 'Premises ID', 'Assignment ID', 'Timestamp'])
        for account_id in processed_accounts:
            sheet3.append([
                account_id,
                'SUCCESS - Created',
                str(account_premises_map.get(account_id, 'N/A')),
                account_assignment_map.get(account_id, 'N/A'),
                datetime.now().isoformat()
            ])
        for account_id in skipped_accounts:
            sheet3.append([
                account_id,
                'SKIPPED - Assignment exists',
                str(account_premises_map.get(account_id, 'N/A')),
                'N/A',
                datetime.now().isoformat()
            ])
        
        # Sheet 4: Schedules
        sheet4: Worksheet = workbook.create_sheet("Schedules")
        sheet4.append(['Account ID', 'Status', 'Premises ID', 'Assignment ID', 'Schedule ID', 'Timestamp'])
        for account_id in processed_accounts:
            sheet4.append([
                account_id,
                'SUCCESS - Created',
                str(account_premises_map.get(account_id, 'N/A')),
                account_assignment_map.get(account_id, 'N/A'),
                account_schedule_map.get(account_id, 'N/A'),
                datetime.now().isoformat()
            ])
        
        # Sheet 5: JSF
        sheet5: Worksheet = workbook.create_sheet("JSF")
        sheet5.append(['Account ID', 'Status', 'Premises ID', 'Assignment ID', 'Schedule ID', 'JSF ID', 'Timestamp'])
        for account_id in processed_accounts:
            sheet5.append([
                account_id,
                'SUCCESS - Created',
                str(account_premises_map.get(account_id, 'N/A')),
                account_assignment_map.get(account_id, 'N/A'),
                account_schedule_map.get(account_id, 'N/A'),
                account_jsf_map.get(account_id, 'N/A'),
                datetime.now().isoformat()
            ])
        
        workbook.save(filename)
        logger.info(f"Excel report generated: {filename}")
        return filename
        
    except Exception as e:
        logger.error(f"Error generating Excel report: {e}")
        return ""

def process_account_assignments(account_ids: List[int]) -> Tuple[int, List[int], List[int], str]:
    """
    Main function to process account assignments with improved logic.
    Returns: (success_count, skipped_accounts, missing_premises_accounts, report_filename)
    """
    connection = None
    try:
        # Establish database connection
        connection = establish_pg_db_connection("sc_ep")
        connection.autocommit = False  # Ensure we're in transaction mode
        
        # Step 1: Validate ALL accounts exist in premises table (stops if any missing)
        account_premises_map = validate_all_accounts_exist(connection, account_ids)
        
        # Step 2: Filter out accounts that already have assignments
        accounts_needing_assignments = filter_accounts_without_assignments(connection, account_premises_map)
        
        # Determine skipped accounts (those with existing assignments)
        skipped_accounts = [acc_id for acc_id in account_ids if acc_id not in accounts_needing_assignments]
        
        # Determine missing premises accounts (those not in premises table)
        missing_premises_accounts = [acc_id for acc_id in account_ids if acc_id not in account_premises_map]
        
        if not accounts_needing_assignments:
            logger.info("No new assignments needed - all accounts either have assignments or are missing premises")
            report_filename = generate_report(account_ids, [], skipped_accounts, missing_premises_accounts, 
                                           account_premises_map, {}, {}, {})
            connection.commit()
            return 0, skipped_accounts, missing_premises_accounts, report_filename
        
        # Step 3: Create assignments (batch)
        account_assignment_map = create_assignments_batch(connection, accounts_needing_assignments)
        
        # Step 4: Create schedules (batch)
        account_schedule_map = create_schedules_batch(connection, account_assignment_map)
        
        # Step 5: Create job satisfaction forms (batch)
        account_jsf_map = create_job_satisfaction_forms_batch(connection, account_schedule_map)
        
        # Commit all transactions
        connection.commit()
        
        # Generate report
        successful_accounts = list(account_assignment_map.keys())
        report_filename = generate_report(
            account_ids,
            successful_accounts, 
            skipped_accounts, 
            missing_premises_accounts,
            account_premises_map, 
            account_assignment_map, 
            account_schedule_map,
            account_jsf_map
        )
        
        logger.info(f"Successfully processed {len(successful_accounts)} accounts, "
                   f"skipped {len(skipped_accounts)} accounts, "
                   f"{len(missing_premises_accounts)} accounts missing premises")
        return len(successful_accounts), skipped_accounts, missing_premises_accounts, report_filename
        
    except AccountProcessingError as e:
        logger.error(f"Account processing stopped: {e}")
        if connection:
            connection.rollback()
        missing_premises_accounts = [acc_id for acc_id in account_ids if acc_id not in validate_all_accounts_exist(connection, account_ids).keys()] if connection else account_ids
        report_filename = generate_report(account_ids, [], [], missing_premises_accounts, {}, {}, {}, {})
        return 0, [], missing_premises_accounts, report_filename
    except Exception as e:
        logger.error(f"Unexpected error in process_account_assignments: {e}")
        if connection:
            connection.rollback()
        return 0, [], account_ids, ""
        
    finally:
        if connection:
            connection.close()

def main(account_ids: Optional[List[int]] = None) -> Tuple[int, List[int], List[int], str]:
    """Main function to execute the assignment creation process."""
    try:
        logger.info("Starting assignment creation process...")
        
        account_ids_to_process = account_ids or ACCOUNT_IDS
        
        if not account_ids_to_process:
            logger.error("No account IDs to process!")
            return 0, [], [], ""
        
        logger.info(f"Processing {len(account_ids_to_process)} account IDs")
        
        # Process assignments
        success_count, skipped_accounts, missing_premises_accounts, report_file = process_account_assignments(account_ids_to_process)
        
        logger.info(f"Process completed:")
        logger.info(f"  - {success_count} new records created")
        logger.info(f"  - {len(skipped_accounts)} accounts skipped (existing assignments)")
        logger.info(f"  - {len(missing_premises_accounts)} accounts missing premises")
        logger.info(f"  - Report file: {report_file}")
        
        return success_count, skipped_accounts, missing_premises_accounts, report_file
        
    except Exception as e:
        logger.error(f"Error in main process: {e}")
        raise

if __name__ == "__main__":
    # Use global variable
    main()