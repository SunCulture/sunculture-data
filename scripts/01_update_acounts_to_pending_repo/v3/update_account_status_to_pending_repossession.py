import psycopg2
from psycopg2.extras import RealDictCursor
import mysql.connector
from dotenv import load_dotenv
import os
import logging
from datetime import datetime
import pandas as pd
from openpyxl import Workbook

# Set up logging
logging.basicConfig(level=logging.WARNING, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()

# Database configurations
DB_CONFIGS = {
    "fma": {
        "host": os.getenv("SC_EP_PG_DB_HOST"),
        "user": os.getenv("SC_EP_PG_DB_USER"),
        "password": os.getenv("SC_EP_PG_DB_PASSWORD"),
        "database": os.getenv("SC_EP_PG_DB_NAME"),
        "port": 5432
    },
    "amtdb_prod": {
        "host": os.getenv("SC_MYSQL_AMTDB_v39_HOST"),
        "user": os.getenv("SC_MYSQL_AMTDB_v39_USER"),
        "password": os.getenv("SC_MYSQL_AMTDB_v39_PASSWORD"),
        "database": os.getenv("SC_MYSQL_AMTDB_v39_DB"),
        "port": 3306
    }
}

# Default user ID for system-generated records (modify as needed)
DEFAULT_USER_ID = 895
REPOSSESSION_REASON_TEMPLATE = "Auto-created from {status} status"
REPOSSESSION_TYPE = "AUTO_GENERATED"
account_ids = ['69031',
'140748',
'111890',
'81436',
'72145',
'81421',
'71189',
'67859',
'125041',
'23011',
'75325',
'129406',
'70178',
'77829',
'138171',
'50347',
'122899',
'122196',
'76490',
'71682',
'76362',
'82611',
'82386',
'84194',
'16209',
'14460',
'66542',
'121110',
'130901',
'106505',
'110914',
'86164',
'80888',
'74302',
'75631',
'77729',
'77257',
'41303',
'80805',
'132856',
'75172',
'69991',
'131778',
'69083',
'122092',
'133530',
'132097',
'84992',
'78547',
'83928',
'125527',
'79215',
'69658',
'137497',
'53800',
'78601',
'141579',
'141687',
'70215',
'81234',
'78613',
'79891',
'71630',
'77369',
'74873',
'142044',
'127726',
'22946',
'91311',
'122826',
'79489',
'140033',
'74732',
'133414',
'138740',
'116745',
'14428',
'67562',
'78789',
'66866',
'75368',
'71325',
'67320',
'82422',
'81323',
'78870',
'125358',
'111018',
'115558']

def get_fma_pg_db_connection():
    """Establish connection to the source FMA PostgreSQL database."""
    try:
        config = DB_CONFIGS["fma"]
        conn = psycopg2.connect(
            host=config["host"],
            user=config["user"],
            password=config["password"],
            database=config["database"],
            port=config["port"]
        )
        print("✅ Connected to FMA database (PostgreSQL)")
        return conn
    except Exception as e:
        logger.error(f"Error connecting to FMA database: {e}")
        raise

def get_amtdb_prod_mysql_db_connection():
    """Establish connection to the target AMTDB_PROD MySQL database."""
    try:
        config = DB_CONFIGS["amtdb_prod"]
        conn = mysql.connector.connect(
            host=config["host"],
            user=config["user"],
            password=config["password"],
            database=config["database"],
            port=config["port"]
        )
        print("✅ Connected to AMTDB_PROD database (MySQL)")
        return conn
    except Exception as e:
        logger.error(f"Error connecting to AMTDB_PROD database: {e}")
        raise

def validate_account_statuses(amtdb_conn, account_ids):
    """
    Validate account statuses from AMTDB_PROD accounts table.
    Returns: tuple of (arrears_accounts, other_accounts, status_counts)
    """
    try:
        cursor = amtdb_conn.cursor(dictionary=True)
        
        # Convert account_ids list to comma-separated string for SQL query
        ids_str = ','.join(map(str, account_ids))
        
        query = f"""
            SELECT id, status 
            FROM accounts 
            WHERE id IN ({ids_str})
        """
        
        cursor.execute(query)
        results = cursor.fetchall()
        cursor.close()
        
        arrears_accounts = []
        other_accounts = []
        status_counts = {}
        
        # Create a set of found account IDs as strings to match account_ids type
        found_ids = set()
        
        # Process accounts found in database
        for row in results:
            account_id = str(row['id'])  # Convert to str to match account_ids
            status = row['status']
            found_ids.add(account_id)
            
            # Count statuses
            status_counts[status] = status_counts.get(status, 0) + 1
            
            if status and status.lower() == 'arrears':
                arrears_accounts.append(account_id)
            else:
                other_accounts.append({
                    'account_id': account_id,
                    'status': status
                })
        
        # Handle accounts not found in database
        missing_accounts = set(account_ids) - found_ids
        if missing_accounts:
            status_counts['NOT FOUND'] = len(missing_accounts)
            for account_id in missing_accounts:
                other_accounts.append({
                    'account_id': account_id,
                    'status': 'NOT FOUND'
                })
        
        print(f"✅ Validated {len(account_ids)} accounts from AMTDB_PROD:")
        print(f"   - Found in database: {len(found_ids)}")
        print(f"   - Arrears status: {len(arrears_accounts)}")
        print(f"   - Other statuses: {len(other_accounts)}")
        print(f"   - Missing accounts: {len(missing_accounts)}")
        
        # Verify the math
        total_counted = sum(status_counts.values())
        print(f"   - Total counted in status_counts: {total_counted}")
        
        return arrears_accounts, other_accounts, status_counts
    
    except Exception as e:
        logger.error(f"Error validating account statuses: {e}")
        raise

def validate_repossession_statuses(fma_conn, arrears_account_ids):
    """
    Validate repossession statuses from FMA repossessions table for arrears accounts.
    Returns: dict of status_counts for repossessions (counting unique accounts, not rows)
    """
    try:
        cursor = fma_conn.cursor(cursor_factory=RealDictCursor)
        
        # Convert arrears_account_ids to comma-separated string for SQL query
        ids_str = ','.join(arrears_account_ids)
        
        # Get the LATEST (most recent) status per account based on created_at timestamp
        # If an account has multiple rows, we only take the one with the highest created_at
        query = f"""
            SELECT DISTINCT ON (account_id) account_id, status, created_at
            FROM repossessions 
            WHERE account_id IN ({ids_str})
            ORDER BY account_id, created_at DESC
        """
        
        cursor.execute(query)
        results = cursor.fetchall()
        cursor.close()
        
        repos_status_counts = {}
        
        # Create a set of found account IDs as strings
        found_ids = set()
        
        # Process accounts found in repossessions (now only one row per account)
        for row in results:
            account_id = str(row['account_id'])
            status = row['status'] or 'NULL'
            found_ids.add(account_id)
            
            # Count statuses
            repos_status_counts[status] = repos_status_counts.get(status, 0) + 1
        
        # Handle accounts not found in repossessions
        missing_accounts = set(arrears_account_ids) - found_ids
        if missing_accounts:
            repos_status_counts['NOT IN REPOSSESSIONS'] = len(missing_accounts)
        
        print(f"✅ Validated repossession statuses for {len(arrears_account_ids)} arrears accounts from FMA:")
        print(f"   - Found in repossessions: {len(found_ids)}")
        print(f"   - Missing in repossessions: {len(missing_accounts)}")
        
        # Verify the math
        total_counted = sum(repos_status_counts.values())
        print(f"   - Total counted in repos_status_counts: {total_counted}")
        
        return repos_status_counts
    
    except Exception as e:
        logger.error(f"Error validating repossession statuses: {e}")
        raise

def create_repossession_records(fma_conn, arrears_account_ids, dry_run=True):
    """
    Create repossession records for arrears accounts based on business rules:
    
    LOGIC 1: Account NOT in repossessions table → CREATE new record with status='NEW'
    LOGIC 2: Account EXISTS with status='NEW' → SKIP (do nothing)
    LOGIC 3: Account EXISTS with status='PAID' or 'CANCELLED' → CREATE new record with status='NEW'
    
    Args:
        fma_conn: PostgreSQL connection to FMA database
        arrears_account_ids: List of account IDs in arrears status
        dry_run: If True, only simulate without actual inserts (default: True)
    
    Returns:
        dict: Summary of actions taken
    """
    try:
        cursor = fma_conn.cursor(cursor_factory=RealDictCursor)
        
        # Convert arrears_account_ids to comma-separated string for SQL query
        ids_str = ','.join(arrears_account_ids)
        
        # Get the LATEST status per account
        query = f"""
            SELECT DISTINCT ON (account_id) account_id, status, created_at
            FROM repossessions 
            WHERE account_id IN ({ids_str})
            ORDER BY account_id, created_at DESC
        """
        
        cursor.execute(query)
        results = cursor.fetchall()
        
        # Build a map of account_id -> latest_status
        existing_repos = {}
        for row in results:
            existing_repos[str(row['account_id'])] = row['status']
        
        # Initialize counters for reporting
        actions = {
            'created_new': [],           # LOGIC 1: Not in repossessions
            'skipped_already_new': [],   # LOGIC 2: Already has status='NEW'
            'created_after_paid': [],    # LOGIC 3: Was PAID, creating new
            'created_after_cancelled': [],  # LOGIC 3: Was CANCELLED, creating new
            'skipped_other_status': []   # Other statuses (not in our 3 logics)
        }
        
        # Process each arrears account
        for account_id in arrears_account_ids:
            if account_id not in existing_repos:
                # LOGIC 1: Account NOT in repossessions table
                actions['created_new'].append(account_id)
            else:
                latest_status = existing_repos[account_id]
                
                if latest_status == 'NEW':
                    # LOGIC 2: Already has status='NEW', skip
                    actions['skipped_already_new'].append(account_id)
                
                elif latest_status == 'PAID':
                    # LOGIC 3: Was PAID, create new record
                    actions['created_after_paid'].append(account_id)
                
                elif latest_status == 'CANCELLED':
                    # LOGIC 3: Was CANCELLED, create new record
                    actions['created_after_cancelled'].append(account_id)
                
                else:
                    # Other statuses - for safety, we skip these
                    actions['skipped_other_status'].append({
                        'account_id': account_id,
                        'status': latest_status
                    })
        
        # Combine all accounts that need new records created
        accounts_to_create = (
            actions['created_new'] + 
            actions['created_after_paid'] + 
            actions['created_after_cancelled']
        )
        
        # INSERT new records if not in dry_run mode
        if not dry_run and accounts_to_create:
            insert_query = """
                INSERT INTO repossessions 
                (account_id, status, created_by, updated_by, flagged_by, flagged_at, reason, type, is_active)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            
            current_time = datetime.now()
            records_inserted = 0
            
            for account_id in accounts_to_create:
                try:
                    # Determine reason based on which list the account is in
                    if account_id in actions['created_new']:
                        reason = "Auto-created from Arrears status (No previous repossession)"
                    elif account_id in actions['created_after_paid']:
                        reason = "Auto-created from Arrears status (Previous status: PAID)"
                    elif account_id in actions['created_after_cancelled']:
                        reason = "Auto-created from Arrears status (Previous status: CANCELLED)"
                    else:
                        reason = "Auto-created from Arrears status"
                    
                    cursor.execute(insert_query, (
                        int(account_id),           # account_id
                        'NEW',                      # status
                        DEFAULT_USER_ID,            # created_by
                        DEFAULT_USER_ID,            # updated_by
                        DEFAULT_USER_ID,            # flagged_by
                        current_time,               # flagged_at
                        reason,                     # reason
                        REPOSSESSION_TYPE,          # type
                        True                        # is_active
                    ))
                    records_inserted += 1
                except Exception as e:
                    logger.error(f"Failed to insert repossession for account {account_id}: {e}")
                    # Continue with other accounts even if one fails
            
            fma_conn.commit()
            print(f"✅ Inserted {records_inserted} new repossession records")
        
        cursor.close()
        
        # Print summary
        print(f"\n{'=' * 60}")
        print(f"REPOSSESSION CREATION SUMMARY {'(DRY RUN)' if dry_run else '(LIVE)'}")
        print(f"{'=' * 60}")
        print(f"Total Arrears Accounts Processed: {len(arrears_account_ids)}")
        print(f"\nLOGIC 1 - Not in repossessions table (CREATE NEW):")
        print(f"  Count: {len(actions['created_new'])}")
        
        print(f"\nLOGIC 2 - Already has status='NEW' (SKIP):")
        print(f"  Count: {len(actions['skipped_already_new'])}")
        
        print(f"\nLOGIC 3a - Previous status='PAID' (CREATE NEW):")
        print(f"  Count: {len(actions['created_after_paid'])}")
        
        print(f"\nLOGIC 3b - Previous status='CANCELLED' (CREATE NEW):")
        print(f"  Count: {len(actions['created_after_cancelled'])}")
        
        print(f"\nOther Status (SKIPPED for safety):")
        print(f"  Count: {len(actions['skipped_other_status'])}")
        if actions['skipped_other_status']:
            for item in actions['skipped_other_status']:
                print(f"    - Account {item['account_id']}: {item['status']}")
        
        print(f"\n{'=' * 60}")
        print(f"TOTAL TO CREATE: {len(accounts_to_create)}")
        print(f"TOTAL TO SKIP: {len(actions['skipped_already_new']) + len(actions['skipped_other_status'])}")
        print(f"{'=' * 60}\n")
        
        return actions
        
    except Exception as e:
        logger.error(f"Error creating repossession records: {e}")
        if not dry_run:
            fma_conn.rollback()
        raise

def update_accounts_to_pending_repossession(amtdb_conn, arrears_account_ids, dry_run=True):
    """
    Update account status to 'Pending Repossession' for arrears accounts that now have 
    repossession records with status='NEW' (either newly created or already existing).
    
    This is the FINAL step after repossession records have been created/validated.
    
    Args:
        amtdb_conn: MySQL connection to AMTDB_PROD database
        arrears_account_ids: List of account IDs currently in 'Arrears' status
        dry_run: If True, only simulate without actual updates (default: True)
    
    Returns:
        dict: Summary of update actions with account details
    """
    try:
        cursor = amtdb_conn.cursor(dictionary=True)
        
        # Convert account_ids to comma-separated string
        ids_str = ','.join(arrears_account_ids)
        
        # First, verify current status of these accounts (safety check)
        verify_query = f"""
            SELECT id, status, updatedAt
            FROM accounts 
            WHERE id IN ({ids_str})
        """
        
        cursor.execute(verify_query)
        current_accounts = cursor.fetchall()
        
        # Build detailed tracking
        update_summary = {
            'accounts_to_update': [],      # Accounts that will be updated
            'accounts_not_arrears': [],    # Safety: accounts not in Arrears status
            'successful_updates': [],      # Successfully updated (live mode only)
            'failed_updates': []           # Failed updates (live mode only)
        }
        
        # Verify each account is actually in 'Arrears' status
        for account in current_accounts:
            account_id = str(account['id'])
            current_status = account['status']
            
            if current_status and current_status.lower() == 'arrears':
                update_summary['accounts_to_update'].append({
                    'account_id': account_id,
                    'current_status': current_status,
                    'new_status': 'Pending Repossession',
                    'previous_updated_at': account['updatedAt']
                })
            else:
                # Safety: Don't update accounts that aren't in Arrears
                update_summary['accounts_not_arrears'].append({
                    'account_id': account_id,
                    'current_status': current_status,
                    'reason': 'Not in Arrears status - skipping for safety'
                })
        
        # Perform updates if not in dry_run mode
        if not dry_run and update_summary['accounts_to_update']:
            update_query = """
                UPDATE accounts 
                SET status = %s, updatedAt = %s
                WHERE id = %s AND status = 'Arrears'
            """
            
            current_time = datetime.now()
            
            for account_info in update_summary['accounts_to_update']:
                account_id = account_info['account_id']
                try:
                    cursor.execute(update_query, (
                        'Pending Repossession',
                        current_time,
                        int(account_id)
                    ))
                    
                    # Check if update was successful (rowcount should be 1)
                    if cursor.rowcount == 1:
                        update_summary['successful_updates'].append(account_info)
                    else:
                        update_summary['failed_updates'].append({
                            **account_info,
                            'reason': f'No rows updated (rowcount={cursor.rowcount})'
                        })
                        
                except Exception as e:
                    update_summary['failed_updates'].append({
                        **account_info,
                        'reason': str(e)
                    })
                    logger.error(f"Failed to update account {account_id}: {e}")
            
            # Commit all updates
            amtdb_conn.commit()
            print(f"✅ Updated {len(update_summary['successful_updates'])} accounts to 'Pending Repossession'")
        
        cursor.close()
        
        # Print summary
        print(f"\n{'=' * 60}")
        print(f"ACCOUNTS STATUS UPDATE SUMMARY {'(DRY RUN)' if dry_run else '(LIVE)'}")
        print(f"{'=' * 60}")
        print(f"Total Arrears Accounts to Process: {len(arrears_account_ids)}")
        print(f"\nAccounts TO UPDATE to 'Pending Repossession':")
        print(f"  Count: {len(update_summary['accounts_to_update'])}")
        
        print(f"\nAccounts SKIPPED (Not in Arrears status):")
        print(f"  Count: {len(update_summary['accounts_not_arrears'])}")
        if update_summary['accounts_not_arrears']:
            for acc in update_summary['accounts_not_arrears']:
                print(f"    - Account {acc['account_id']}: Current status = '{acc['current_status']}'")
        
        if not dry_run:
            print(f"\n✅ SUCCESSFULLY UPDATED:")
            print(f"  Count: {len(update_summary['successful_updates'])}")
            
            if update_summary['failed_updates']:
                print(f"\n❌ FAILED UPDATES:")
                print(f"  Count: {len(update_summary['failed_updates'])}")
                for acc in update_summary['failed_updates']:
                    print(f"    - Account {acc['account_id']}: {acc['reason']}")
        
        print(f"{'=' * 60}\n")
        
        return update_summary
        
    except Exception as e:
        logger.error(f"Error updating accounts status: {e}")
        if not dry_run:
            amtdb_conn.rollback()
        raise

def create_enhanced_excel_report(account_ids, status_counts, arrears_count, repos_status_counts, 
                                  repos_actions, update_summary=None, dry_run=True, output_file="account_status_report.xlsx"):
    """
    Create an Excel report showing:
    1. Account status counts
    2. Repossession status for arrears accounts
    3. Actions to be taken on repossession records (NEW LAYER)
    """
    try:
        # Create a new workbook and select the active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Account Status Report"
        
        # === SECTION 1: Account Status Summary ===
        ws['A1'] = "Accounts to be Updated to Pending Repossession"
        ws['A2'] = "Generated Datetime"
        ws['B2'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws['A3'] = "Mode"
        ws['B3'] = "DRY RUN (No Changes)" if dry_run else "LIVE (Changes Applied)"
        ws['A4'] = "Total Accounts"
        ws['B4'] = len(account_ids)
        
        current_row = 6
        ws[f'A{current_row}'] = "Status"
        ws[f'B{current_row}'] = "Count"
        current_row += 1
        
        # Write account status counts
        for status, count in sorted(status_counts.items()):
            ws[f'A{current_row}'] = status
            ws[f'B{current_row}'] = count
            current_row += 1
        
        # === SECTION 2: Repossession Status for Arrears Accounts ===
        current_row += 2
        ws[f'A{current_row}'] = "Repossession Status for Arrears Accounts"
        current_row += 2
        ws[f'A{current_row}'] = "Total Arrears Accounts"
        ws[f'B{current_row}'] = arrears_count
        current_row += 2
        
        ws[f'A{current_row}'] = "Status"
        ws[f'B{current_row}'] = "Count"
        current_row += 1
        
        # Write repossession status counts
        for status, count in sorted(repos_status_counts.items()):
            ws[f'A{current_row}'] = status
            ws[f'B{current_row}'] = count
            current_row += 1
        
        # === SECTION 3: Repossession Actions (NEW) ===
        current_row += 2
        ws[f'A{current_row}'] = "Repossession Record Actions"
        ws[f'A{current_row}'].font = ws[f'A{current_row}'].font.copy(bold=True)
        current_row += 2
        
        ws[f'A{current_row}'] = "Action Type"
        ws[f'B{current_row}'] = "Count"
        ws[f'C{current_row}'] = "Description"
        current_row += 1
        
        # LOGIC 1
        ws[f'A{current_row}'] = "CREATE NEW"
        ws[f'B{current_row}'] = len(repos_actions['created_new'])
        ws[f'C{current_row}'] = "Not in repossessions table - creating NEW record"
        current_row += 1
        
        # LOGIC 2
        ws[f'A{current_row}'] = "SKIP (Already NEW)"
        ws[f'B{current_row}'] = len(repos_actions['skipped_already_new'])
        ws[f'C{current_row}'] = "Already has status='NEW' - no action needed"
        current_row += 1
        
        # LOGIC 3a
        ws[f'A{current_row}'] = "CREATE NEW (After PAID)"
        ws[f'B{current_row}'] = len(repos_actions['created_after_paid'])
        ws[f'C{current_row}'] = "Previous status='PAID' - creating NEW record"
        current_row += 1
        
        # LOGIC 3b
        ws[f'A{current_row}'] = "CREATE NEW (After CANCELLED)"
        ws[f'B{current_row}'] = len(repos_actions['created_after_cancelled'])
        ws[f'C{current_row}'] = "Previous status='CANCELLED' - creating NEW record"
        current_row += 1
        
        # Other statuses skipped
        ws[f'A{current_row}'] = "SKIP (Other Status)"
        ws[f'B{current_row}'] = len(repos_actions['skipped_other_status'])
        ws[f'C{current_row}'] = "Other status - skipped for safety"
        current_row += 1
        
        # Totals
        current_row += 1
        total_to_create = (len(repos_actions['created_new']) + 
                          len(repos_actions['created_after_paid']) + 
                          len(repos_actions['created_after_cancelled']))
        total_to_skip = (len(repos_actions['skipped_already_new']) + 
                        len(repos_actions['skipped_other_status']))
        
        ws[f'A{current_row}'] = "TOTAL TO CREATE"
        ws[f'B{current_row}'] = total_to_create
        ws[f'A{current_row}'].font = ws[f'A{current_row}'].font.copy(bold=True)
        current_row += 1
        
        ws[f'A{current_row}'] = "TOTAL TO SKIP"
        ws[f'B{current_row}'] = total_to_skip
        ws[f'A{current_row}'].font = ws[f'A{current_row}'].font.copy(bold=True)
        current_row += 1
        
        # Validation check
        current_row += 1
        ws[f'A{current_row}'] = "Validation"
        ws[f'B{current_row}'] = "PASS" if (total_to_create + total_to_skip == arrears_count) else "FAIL"
        ws[f'C{current_row}'] = f"Total actions ({total_to_create + total_to_skip}) should equal arrears count ({arrears_count})"
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 80)  # Cap at 80 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        wb.save(output_file)
        print(f"✅ Excel report generated: {output_file}")
        
        # Console summary
        print(f"\n📊 ENHANCED REPORT SUMMARY:")
        print(f"   Mode: {'DRY RUN' if dry_run else 'LIVE'}")
        print(f"   Total Accounts: {len(account_ids)}")
        print(f"   Arrears Accounts: {arrears_count}")
        print(f"   Actions Summary:")
        print(f"     - Records to CREATE: {total_to_create}")
        print(f"     - Records to SKIP: {total_to_skip}")
        print(f"   Validation: {'✅ PASS' if (total_to_create + total_to_skip == arrears_count) else '❌ FAIL'}")
            
        return output_file
        
    except Exception as e:
        logger.error(f"Error creating enhanced Excel report: {e}")
        raise

    
# Updated main function
def main():
    """
    Main workflow:
    1. Validate account statuses from AMTDB_PROD
    2. Validate repossession statuses from FMA
    3. Apply business logic to create/skip repossession records
    4. Generate comprehensive Excel report
    
    Usage:
        py script.py              # Dry run (default)
        py script.py --dry-run    # Dry run (explicit)
        py script.py --execute    # Live execution
    """
    import sys
    
    # Default to dry run for safety
    DRY_RUN = False
    
    # Check command-line arguments
    if len(sys.argv) > 1:
        if '--execute' in sys.argv or '--live' in sys.argv:
            DRY_RUN = False
            confirmation = input("\n⚠️  WARNING: You are about to make LIVE changes to the database.\nType 'YES' to confirm: ")
            if confirmation != 'YES':
                print("❌ Execution cancelled. Exiting.")
                return
        elif '--dry-run' in sys.argv:
            DRY_RUN = True
    
    try:
        global account_ids
        
        if not account_ids:
            print("❌ No account IDs provided. Please populate the account_ids list.")
            return
        
        print(f"\n{'='*60}")
        print(f"STARTING REPOSSESSION RECORD CREATION PROCESS")
        print(f"Mode: {'DRY RUN (No changes will be made)' if DRY_RUN else 'LIVE (Changes will be applied)'}")
        print(f"{'='*60}\n")
        
        # STEP 1: Connect to AMTDB_PROD and validate account statuses
        print("STEP 1: Validating account statuses from AMTDB_PROD...")
        amtdb_conn = get_amtdb_prod_mysql_db_connection()
        arrears_accounts, non_arrears_accounts, status_counts = validate_account_statuses(amtdb_conn, account_ids)
        amtdb_conn.close()
        
        if not arrears_accounts:
            print("❌ No arrears accounts found. Exiting.")
            return
        
        # STEP 2: Connect to FMA and validate repossession statuses
        print("\nSTEP 2: Validating repossession statuses from FMA...")
        fma_conn = get_fma_pg_db_connection()
        repos_status_counts = validate_repossession_statuses(fma_conn, arrears_accounts)
        
        # STEP 3: Apply business logic to create/skip records
        print("\nSTEP 3: Applying business logic for repossession records...")
        repos_actions = create_repossession_records(fma_conn, arrears_accounts, dry_run=DRY_RUN)
        
        # Close FMA connection
        fma_conn.close()
        
        # STEP 4: Update accounts status to 'Pending Repossession'
        print("\nSTEP 4: Updating accounts status to 'Pending Repossession'...")
        amtdb_conn = get_amtdb_prod_mysql_db_connection()
        update_summary = update_accounts_to_pending_repossession(amtdb_conn, arrears_accounts, dry_run=DRY_RUN)
        amtdb_conn.close()
        
        # STEP 5: Generate comprehensive Excel report
        print("\nSTEP 5: Generating Excel report...")
        arrears_count = len(arrears_accounts)
        report_filename = create_enhanced_excel_report(
            account_ids, 
            status_counts,
            arrears_count,
            repos_status_counts,
            repos_actions,
            update_summary,
            dry_run=DRY_RUN,
            output_file=f"repossession_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        print(f"\n{'='*60}")
        print(f"🎯 Process completed successfully!")
        print(f"{'='*60}")
        print(f"Report saved as: {report_filename}")
        
        if DRY_RUN:
            print(f"\n⚠️  This was a DRY RUN - No changes were made to the database")
            print(f"   To apply changes, set DRY_RUN = False in the script")
        else:
            print(f"\n✅ Changes have been applied to the database")
        
        print(f"{'='*60}\n")
        
    except Exception as e:
        logger.error(f"Error in main execution: {e}")
        raise

if __name__ == "__main__":
    main()